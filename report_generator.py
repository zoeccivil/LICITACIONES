# report_generator.py

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.lib import colors
from tkinter import messagebox
from reportlab.lib.pagesizes import letter, landscape, legal, elevenSeventeen
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
from reportlab.lib import colors
from datetime import datetime
OPENPYXL_AVAILABLE = True
REPORTLAB_AVAILABLE = True

class ReportGenerator:
    """
    Clase dedicada a la generación de todos los reportes de la aplicación,
    tanto en formato Excel (.xlsx) como en PDF.
    """
    # ======= COLORES DE MARCA / ESTILOS =======
    GREEN_DARK  = colors.Color(0/255, 99/255, 65/255)     # encabezado tabla
    GREEN_LIGHT = colors.Color(209/255, 242/255, 223/255) # ganador
    ROW_STRIPE  = colors.Color(245/255, 245/255, 245/255) # alternado

    def _header_footer(self, canvas, doc, licitacion):
        """Encabezado y pie de página en todas las páginas."""
        canvas.saveState()

        # Encabezado
        y_top = doc.pagesize[1] - 0.4*inch
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawString(doc.leftMargin, y_top, "Reporte de Evaluación de Ofertas")
        canvas.setFont("Helvetica", 9)
        linea2 = f"Licitación: {licitacion.nombre_proceso}  ({licitacion.numero_proceso})   •   Institución: {licitacion.institucion}"
        canvas.drawString(doc.leftMargin, y_top - 12, linea2)

        # Pie de página
        canvas.setFont("Helvetica", 8)
        fecha_txt = datetime.now().strftime("%d/%m/%Y %H:%M")
        pie_izq = f"Generado: {fecha_txt}"
        pie_der = f"Página {doc.page}"
        canvas.drawString(doc.leftMargin, 0.3*inch, pie_izq)
        canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, 0.3*inch, pie_der)

        canvas.restoreState()


    # --------------------- API PÚBLICA ---------------------
    def generate_bid_results_report(self, licitacion, file_path):
        """
        Genera el reporte de resultados para una licitación específica.

        Args:
            licitacion (Licitacion): El objeto de la licitación a reportar.
            file_path (str): La ruta donde se guardará el archivo (.xlsx o .pdf).
        """
        if file_path.endswith('.xlsx'):
            if not OPENPYXL_AVAILABLE:
                messagebox.showerror("Librería Faltante", "La librería 'openpyxl' es necesaria para exportar a Excel. Instala: pip install openpyxl")
                return
            self._generate_bid_excel(licitacion, file_path)
        elif file_path.endswith('.pdf'):
            if not REPORTLAB_AVAILABLE:
                messagebox.showerror("Librería Faltante", "La librería 'reportlab' es necesaria para exportar a PDF. Instala: pip install reportlab")
                return
            self._generate_bid_pdf(licitacion=licitacion, file_path=file_path)

    def generate_institution_history_report(self, all_bids, file_path):
        """
        Genera un reporte histórico de todas las licitaciones, agrupadas o no por institución.

        Args:
            all_bids (list or dict): Una lista de objetos Licitacion o un diccionario.
            file_path (str): La ruta donde se guardará el archivo.
        """
        if file_path.endswith('.xlsx'):
            if not OPENPYXL_AVAILABLE:
                messagebox.showerror("Librería Faltante", "Se necesita 'openpyxl' para exportar a Excel.")
                return
            self._generate_institution_excel(all_bids, file_path)
        elif file_path.endswith('.pdf'):
            if not REPORTLAB_AVAILABLE:
                messagebox.showerror("Librería Faltante", "Se necesita 'reportlab' para exportar a PDF.")
                return
            self._generate_institution_pdf(all_bids, file_path)


    # --------------------- HELPERS INTERNOS ---------------------
    @staticmethod
    def _norm(s):
        """Normaliza un string a mayúsculas y sin espacios extra para comparación."""
        s = (s or "").strip()
        s = s.replace("➡️", "").replace("(Nuestra Oferta)", "")
        while "  " in s:
            s = s.replace("  ", " ")
        return s.upper()

    def _map_ganadores_por_lote(self, lic):
        """Devuelve un mapa { 'num_lote': {'ganador': str, 'es_nuestro': bool} }."""
        winners = {}
        for lote in getattr(lic, "lotes", []):
            winners[str(getattr(lote, "numero", ""))] = {
                "ganador": (getattr(lote, "ganador_nombre", "") or "").strip(),
                "es_nuestro": bool(getattr(lote, "ganado_por_nosotros", False)),
                "empresa_nuestra": getattr(lote, "empresa_nuestra", None)
            }
        return winners

    # --------------------- EXCEL RESULTADOS ---------------------
    def _generate_bid_excel(self, licitacion, file_path):
        wb = openpyxl.Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"

        # Estilos
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True)
        bold_font = Font(bold=True)
        winner_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        winner_ours_fill = PatternFill(start_color="B7E1A1", end_color="B7E1A1", fill_type="solid")
        red_italic = Font(italic=True, color="FF0000")

        # Resumen
        ws_resumen['A1'] = f"Reporte de Licitación: {licitacion.nombre_proceso}"
        ws_resumen['A1'].font = header_font
        resumen = [
            ("Código Proceso", licitacion.numero_proceso),
            ("Institución", licitacion.institucion),
            ("Nuestras Empresas", ", ".join(str(e) for e in licitacion.empresas_nuestras)),
            ("Estado Actual", licitacion.estado),
            ("Monto Base Total", f"RD$ {licitacion.get_monto_base_total():,.2f}"),
            ("Monto Ofertado Total", f"RD$ {licitacion.get_oferta_total():,.2f}"),
            ("Diferencia (%)", f"{licitacion.get_diferencia_porcentual():.2f}%"),
            ("Progreso Docs", f"{licitacion.get_porcentaje_completado():.1f}%"),
        ]
        r = 3
        for k, v in resumen:
            ws_resumen.cell(row=r, column=1, value=k).font = title_font
            ws_resumen.cell(row=r, column=2, value=v)
            r += 1

        # Resultados Competidores
        ws = wb.create_sheet("Resultados Competidores")
        ws.append(["Participante / Lote", "Monto Ofertado", "Monto Habilitado (A)", "Estado A", "Base Lote", "% Dif.", "Ganador", "Empresa Nuestra"])
        for c in ws[1]:
            c.font = bold_font
            c.alignment = Alignment(horizontal='center')

        winners_by_lot = self._map_ganadores_por_lote(licitacion)
        nuestras_empresas = {self._norm(str(e)) for e in getattr(licitacion, "empresas_nuestras", [])}

        participantes = [o.__dict__ for o in getattr(licitacion, "oferentes_participantes", [])]
        nuestras = ", ".join(str(e) for e in getattr(licitacion, "empresas_nuestras", [])) or "Nuestras Empresas"
        nuestras_ofertas = [
            {'lote_numero': l.numero, 'monto': l.monto_ofertado, 'paso_fase_A': l.fase_A_superada}
            for l in getattr(licitacion, "lotes", [])
            if getattr(l, "participamos", False)
        ]
        participantes.append({"nombre": f"➡️ {nuestras} (Nuestra Oferta)", "es_nuestra": True, "ofertas_por_lote": nuestras_ofertas})

        def _monto_hab_total(p):
            return sum(o.get('monto', 0) for o in p.get("ofertas_por_lote", []) if o.get('paso_fase_A', False))

        participantes_orden = sorted(
            participantes,
            key=lambda it: _monto_hab_total(it) if _monto_hab_total(it) > 0 else float('inf')
        )

        for p in participantes_orden:
            nombre = p.get("nombre", "")
            fila_padre = [nombre, "", f"RD$ {_monto_hab_total(p):,.2f}" if _monto_hab_total(p) > 0 else "N/D", "", "", "", "", ""]
            ws.append(fila_padre)
            row_padre = ws.max_row
            ws.cell(row=row_padre, column=1).font = bold_font
            gano_alguno = 0

            for oferta in sorted(p.get("ofertas_por_lote", []), key=lambda o: str(o.get('lote_numero', ''))):
                num = str(oferta.get('lote_numero', ''))
                lot = next((l for l in getattr(licitacion, "lotes", []) if str(l.numero) == num), None)
                nombre_lote = getattr(lot, "nombre", "N/E")
                base = float(getattr(lot, "monto_base", 0) or 0)
                monto = float(oferta.get('monto', 0) or 0)
                pasoA = bool(oferta.get('paso_fase_A', False))

                dif = ""
                if base > 0 and monto > 0:
                    dif = f"{((monto - base)/base)*100:.2f}%"

                info_g = winners_by_lot.get(num, {})
                ganador_real = self._norm(info_g.get("ganador", ""))
                es_nuestro_ganador = bool(info_g.get("es_nuestro", False))

                nombres_en_padre = {x.strip() for x in self._norm(nombre).split(",") if x.strip()}
                es_ganador_esta_fila = False
                if ganador_real:
                    if p.get("es_nuestra") and (ganador_real in nuestras_empresas):
                        es_ganador_esta_fila = True
                    elif ganador_real in nombres_en_padre:
                        es_ganador_esta_fila = True
                    elif self._norm(nombre).startswith(ganador_real):
                        es_ganador_esta_fila = True

                fila = [
                    f"  ↳ Lote {num}: {nombre_lote}",
                    f"RD$ {monto:,.2f}" if monto > 0 else "RD$ 0.00",
                    "",
                    "✅" if pasoA else "❌",
                    f"RD$ {base:,.2f}" if base > 0 else "N/D",
                    dif or "N/D",
                    "Sí" if es_ganador_esta_fila else "No",
                    info_g.get("empresa_nuestra") if es_nuestro_ganador else ""
                ]
                ws.append(fila)
                row = ws.max_row

                if not pasoA:
                    for c in ws[row]:
                        c.font = red_italic
                if es_ganador_esta_fila:
                    for c in ws[row]:
                        c.fill = winner_ours_fill if (p.get("es_nuestra") and es_nuestro_ganador) else winner_fill
                    gano_alguno += 1

            if gano_alguno > 0:
                for c in ws[row_padre]:
                    c.fill = winner_ours_fill if p.get("es_nuestra") else winner_fill
                ws.cell(row=row_padre, column=7, value=f"Sí ({gano_alguno})")

        for col in ws.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                v = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(v))
            ws.column_dimensions[letter].width = max_len + 2

        wb.save(file_path)

    # --------------------- PDF RESULTADOS ---------------------
    def _generate_bid_pdf(self, licitacion, file_path):
        lic = licitacion
        doc = SimpleDocTemplate(
            file_path,
            pagesize=landscape(letter),
            leftMargin=0.5*inch, rightMargin=0.5*inch,
            topMargin=0.5*inch, bottomMargin=0.5*inch
        )

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="small", fontSize=8, leading=10, wordWrap='CJK', alignment=TA_LEFT))
        styles.add(ParagraphStyle(name="small_right", fontSize=8, leading=10, wordWrap='CJK', alignment=TA_RIGHT))
        styles.add(ParagraphStyle(name="small_center", fontSize=8, leading=10, wordWrap='CJK', alignment=TA_CENTER))
        styles.add(ParagraphStyle(name="hsmall", fontSize=9, leading=11, wordWrap='CJK', alignment=TA_LEFT))

        elems = [Paragraph("Resultados Detallados", styles["h1"]), Paragraph(lic.nombre_proceso, styles["h2"]), Spacer(1, 0.15*inch)]
        
        head = ["Participante / Lote", "Monto Ofertado", "Habilitado (A)", "Estado A", "Base Lote", "% Dif.", "Ganador", "Empresa Nuestra"]
        data = [[Paragraph(h, styles["small_center"]) for h in head]]
        tstyle = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#DDDDDD")),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.6, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ALIGN', (1,1), (1,-1), 'RIGHT'), ('ALIGN', (2,1), (2,-1), 'RIGHT'),
            ('ALIGN', (3,1), (3,-1), 'CENTER'), ('ALIGN', (4,1), (4,-1), 'RIGHT'),
            ('ALIGN', (5,1), (5,-1), 'RIGHT'), ('ALIGN', (6,1), (6,-1), 'CENTER'),
            ('ALIGN', (7,1), (7,-1), 'LEFT'),
        ])

        winners_by_lot = self._map_ganadores_por_lote(lic)
        nuestras_empresas = {self._norm(str(e)) for e in getattr(lic, "empresas_nuestras", [])}

        participantes = [o.__dict__ for o in getattr(lic, "oferentes_participantes", [])]
        nuestras = ", ".join(str(e) for e in getattr(lic, "empresas_nuestras", [])) or "Nuestras Empresas"
        nuestras_ofertas = [
            {'lote_numero': l.numero, 'monto': l.monto_ofertado, 'paso_fase_A': l.fase_A_superada}
            for l in getattr(lic, "lotes", [])
            if getattr(l, "participamos", False)
        ]
        participantes.append({"nombre": f"➡️ {nuestras} (Nuestra Oferta)", "es_nuestra": True, "ofertas_por_lote": nuestras_ofertas})

        def _monto_hab_total(p):
            return sum(o.get('monto', 0) for o in p.get("ofertas_por_lote", []) if o.get('paso_fase_A', False))

        participantes_orden = sorted(participantes, key=lambda it: _monto_hab_total(it) if _monto_hab_total(it) > 0 else float('inf'))

        current = 1
        for p in participantes_orden:
            nombre = p.get("nombre", "")
            padre = [
                Paragraph(f"<b>{nombre}</b>", styles["hsmall"]), Paragraph("", styles["small_right"]),
                Paragraph(f"RD$ {_monto_hab_total(p):,.2f}" if _monto_hab_total(p) > 0 else "N/D", styles["small_right"]),
                Paragraph("", styles["small_center"]), Paragraph("", styles["small_right"]),
                Paragraph("", styles["small_right"]), Paragraph("", styles["small_center"]), Paragraph("", styles["small"])
            ]
            data.append(padre)
            row_padre = current
            current += 1

            gano_alguno = 0
            for oferta in sorted(p.get("ofertas_por_lote", []), key=lambda o: str(o.get('lote_numero', ''))):
                num = str(oferta.get('lote_numero', ''))
                lot = next((l for l in getattr(lic, "lotes", []) if str(l.numero) == num), None)
                nombre_lote = getattr(lot, "nombre", "N/E")
                base = float(getattr(lot, "monto_base", 0) or 0)
                monto = float(oferta.get("monto", 0) or 0)
                pasoA = bool(oferta.get("paso_fase_A", False))

                dif = f"{((monto - base)/base)*100:.2f}%" if base > 0 and monto > 0 else ""
                
                info_g = winners_by_lot.get(num, {})
                ganador_real_norm = self._norm(info_g.get("ganador", ""))
                es_nuestro_ganador = bool(info_g.get("es_nuestro", False))

                nombres_en_padre = {x.strip() for x in self._norm(nombre).split(",") if x.strip()}
                es_ganador_esta_fila = False
                if ganador_real_norm:
                    if p.get("es_nuestra") and (ganador_real_norm in nuestras_empresas):
                        es_ganador_esta_fila = True
                    elif ganador_real_norm in nombres_en_padre:
                        es_ganador_esta_fila = True
                    elif self._norm(nombre).startswith(ganador_real_norm):
                        es_ganador_esta_fila = True

                fila = [
                    Paragraph(f"↳ Lote {num}: {nombre_lote}", styles["small"]), Paragraph(f"RD$ {monto:,.2f}", styles["small_right"]),
                    Paragraph("", styles["small_right"]), Paragraph("✅" if pasoA else "❌", styles["small_center"]),
                    Paragraph(f"RD$ {base:,.2f}" if base > 0 else "N/D", styles["small_right"]), Paragraph(dif or "N/D", styles["small_right"]),
                    Paragraph("Sí" if es_ganador_esta_fila else "No", styles["small_center"]), Paragraph(info_g.get("empresa_nuestra") if es_nuestro_ganador else "", styles["small"])
                ]
                data.append(fila)

                if not pasoA:
                    tstyle.add('TEXTCOLOR', (0, current), (-1, current), colors.red)
                if es_ganador_esta_fila:
                    tstyle.add('BACKGROUND', (0, current), (-1, current), colors.lightgreen)
                    gano_alguno += 1
                current += 1

            if gano_alguno > 0:
                tstyle.add('BACKGROUND', (0, row_padre), (-1, row_padre), colors.lightgreen)
                data[row_padre][6] = Paragraph(f"Sí ({gano_alguno})", styles["small_center"])

        ratios = [0.30, 0.13, 0.12, 0.08, 0.12, 0.10, 0.08, 0.07]
        col_widths = [doc.width * r for r in ratios]
        table = Table(data, colWidths=col_widths, repeatRows=1, splitByRow=True)
        table.setStyle(tstyle)
        elems.append(table)
        doc.build(elems)

    # --------------------- EXCEL/PDF HISTÓRICO POR INSTITUCIÓN ---------------------
    def _generate_institution_excel(self, all_bids, file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Histórico por Institución"

        headers = ["Institución", "Proceso", "Nuestras Empresas", "Monto Ofertado Total", "Estado", "Fase A Habilitada", "Comentarios", "Empresa Nuestra Adjudicada"]
        ws.append(headers)
        for c in ws[1]: c.font = Font(bold=True)

        def _append_row(lic):
            empresas_str = ", ".join(str(e) for e in lic.empresas_nuestras)
            habilitado_str = "Sí" if getattr(lic, "fase_A_superada", False) else "No"
            adjudicada_ntra = ""
            if getattr(lic, "estado", "") == "Adjudicada":
                if getattr(lic, "adjudicada_a", "") in {str(e) for e in lic.empresas_nuestras}:
                    adjudicada_ntra = lic.adjudicada_a
            ws.append([getattr(lic, "institucion", ""), lic.nombre_proceso, empresas_str,
                       lic.get_oferta_total(), lic.estado, habilitado_str,
                       getattr(lic, "motivo_descalificacion", ""), adjudicada_ntra])

        if isinstance(all_bids, dict):
            for _, lst in all_bids.items():
                for lic in lst: _append_row(lic)
        else:
            for lic in all_bids: _append_row(lic)

        for col in ws.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                v = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(v))
            ws.column_dimensions[letter].width = max_len + 2

        wb.save(file_path)

# Pega estos 3 nuevos métodos completos dentro de tu clase ReportGenerator

    def generate_package_analysis_report(self, licitacion, file_path):
        """
        Genera un reporte del análisis de paquetes (tabla pivote y resumen).
        """
        if file_path.endswith('.xlsx'):
            if not OPENPYXL_AVAILABLE:
                messagebox.showerror("Librería Faltante", "Se necesita 'openpyxl' para exportar a Excel.")
                return
            self._generate_package_excel(licitacion, file_path)
        elif file_path.endswith('.pdf'):
            if not REPORTLAB_AVAILABLE:
                messagebox.showerror("Librería Faltante", "Se necesita 'reportlab' para exportar a PDF.")
                return
            self._generate_package_pdf(licitacion, file_path)

    def _generate_package_excel(self, licitacion, file_path):
        """Genera el reporte de análisis de paquetes en formato Excel."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analisis de Paquetes"

        # Estilos
        header_font = Font(bold=True, size=12)
        winner_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        
        # Título
        ws['A1'] = f"Análisis de Paquetes: {licitacion.nombre_proceso}"
        ws['A1'].font = Font(bold=True, size=14)

        # Tabla Pivote
        ws.merge_cells('A3:D3')
        ws['A3'] = "Tabla Pivote de Ofertas"
        ws['A3'].font = header_font
        
        matriz = licitacion.get_matriz_ofertas()
        oferentes = sorted(list(set(o for ofertas in matriz.values() for o in ofertas)))
        
        headers = ["Lote"] + oferentes
        ws.append(headers)
        
        start_row = ws.max_row
        for lote_num, ofertas in sorted(matriz.items()):
            lote_obj = next((l for l in licitacion.lotes if str(l.numero) == lote_num), None)
            nombre_lote = lote_obj.nombre if lote_obj else 'N/D'
            
            valores_fila = [f"Lote {lote_num}: {nombre_lote}"]
            for oferente in oferentes:
                oferta = ofertas.get(oferente)
                valores_fila.append(oferta['monto'] if oferta and isinstance(oferta.get('monto'), (int, float)) else "")
            ws.append(valores_fila)
        
        # Resaltar la mejor oferta por lote
        for row_idx in range(start_row + 1, ws.max_row + 1):
            montos = [cell.value for cell in ws[row_idx][1:] if isinstance(cell.value, (int, float))]
            if not montos: continue
            min_monto = min(montos)
            for cell in ws[row_idx][1:]:
                if cell.value == min_monto:
                    cell.fill = winner_fill
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '"RD$"#,##0.00'

        # Resumen del Análisis
        row_idx = ws.max_row + 3
        ws[f'A{row_idx}'] = "Resultados del Análisis"
        ws[f'A{row_idx}'].font = header_font
        row_idx += 1
        
        paquete_individual = licitacion.calcular_mejor_paquete_individual()
        paquete_unico = licitacion.calcular_mejor_paquete_por_oferente()

        ws[f'A{row_idx}'] = "Opción 1: Mejor Oferta por Lote Individual"
        ws[f'B{row_idx}'] = paquete_individual['monto_total']
        ws[f'B{row_idx}'].number_format = '"RD$"#,##0.00'
        row_idx += 1

        ws[f'A{row_idx}'] = "Opción 2: Mejor Paquete de Oferente Único"
        if paquete_unico:
            ws[f'B{row_idx}'] = paquete_unico['monto_total']
            ws[f'C{row_idx}'] = f"({paquete_unico['oferente']})"
            ws[f'B{row_idx}'].number_format = '"RD$"#,##0.00'
        else:
            ws[f'B{row_idx}'] = "N/A"
        
        for col in ws.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        wb.save(file_path)

    def _generate_package_pdf(self, licitacion, file_path):
            """
            Genera el reporte de análisis de paquetes en formato PDF, tamaño 11x17" y horizontal,
            con análisis detallado por lote.
            """
            # --- CAMBIO 1: Usar tamaño 11x17 (ledger) en landscape ---
            doc = SimpleDocTemplate(file_path, pagesize=landscape(elevenSeventeen),
                                    leftMargin=0.5*inch, rightMargin=0.5*inch,
                                    topMargin=0.5*inch, bottomMargin=0.5*inch)
            
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(name="small", fontSize=8, leading=10))
            styles.add(ParagraphStyle(name="small_right", fontSize=8, leading=10, alignment=TA_RIGHT))
            styles.add(ParagraphStyle(name="h2_left", parent=styles['h2'], alignment=TA_LEFT))
            elems = []

            # --- Títulos ---
            elems.append(Paragraph(f"Análisis de Paquetes de Ofertas", styles['h1']))
            elems.append(Paragraph(f"Licitación: {licitacion.nombre_proceso}", styles['h2']))
            elems.append(Spacer(1, 0.2*inch))

            # --- CAMBIO 2: Incluir nuestra oferta en la matriz de datos ---
            matriz = licitacion.get_matriz_ofertas()
            for lote in licitacion.lotes:
                if getattr(lote, 'participamos', False) and float(getattr(lote, 'monto_ofertado', 0) or 0) > 0:
                    lote_num_str = str(lote.numero)
                    # Usamos un nombre distintivo para nuestra empresa
                    empresa_nuestra = f"➡️ {lote.empresa_nuestra or 'Nuestra Oferta'}"
                    matriz.setdefault(lote_num_str, {})[empresa_nuestra] = {'monto': lote.monto_ofertado}

            if not matriz:
                elems.append(Paragraph("No hay ofertas habilitadas para analizar.", styles['Normal']))
                doc.build(elems)
                return

            # --- Tabla Pivote (ahora incluye nuestra oferta) ---
            elems.append(Paragraph("Tabla Comparativa de Ofertas", styles['h3']))
            oferentes = sorted(list(set(o for ofertas in matriz.values() for o in ofertas)))
            
            header = [Paragraph(f"<b>Lote</b>", styles['small'])] + [Paragraph(f"<b>{o}</b>", styles['small']) for o in oferentes]
            data = [header]
            
            tstyle = TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.grey), ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('GRID', (0,0), (-1,-1), 1, colors.black), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (1,1), (-1,-1), 'RIGHT')
            ])
            
            for row_idx, (lote_num, ofertas) in enumerate(sorted(matriz.items()), start=1):
                lote_obj = next((l for l in licitacion.lotes if str(l.numero) == lote_num), None)
                nombre_lote_completo = f"Lote {lote_num}: {lote_obj.nombre if lote_obj else ''}"
                valores_fila = [Paragraph(nombre_lote_completo, styles['small'])]
                montos = [d['monto'] for d in ofertas.values() if isinstance(d.get('monto'), (int, float))]
                min_monto = min(montos) if montos else None

                for col_idx, oferente in enumerate(oferentes, start=1):
                    oferta = ofertas.get(oferente)
                    if oferta and isinstance(oferta.get('monto'), (int, float)):
                        monto = oferta['monto']
                        cell_text = f"RD$ {monto:,.2f}"
                        valores_fila.append(Paragraph(cell_text, styles['small_right']))
                        if monto == min_monto:
                            tstyle.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.lightgreen)
                    else:
                        valores_fila.append(Paragraph("---", styles['small_right']))
                data.append(valores_fila)
            
            ancho_util = doc.width
            ancho_col_lote = ancho_util * 0.20
            ancho_col_oferente = (ancho_util * 0.80) / len(oferentes) if oferentes else 0
            col_widths = [ancho_col_lote] + [ancho_col_oferente] * len(oferentes)

            table = Table(data, colWidths=col_widths, hAlign='LEFT', repeatRows=1)
            table.setStyle(tstyle)
            elems.append(table)
            elems.append(Spacer(1, 0.3*inch))

            # --- CAMBIO 3: Nuevo Resumen y Análisis ---
            elems.append(Paragraph("Análisis de Ofertas Más Bajas por Lote", styles['h2_left']))
            
            analisis_por_lote = []
            for lote_num, ofertas_lote in sorted(matriz.items()):
                lote_obj = next((l for l in licitacion.lotes if str(l.numero) == lote_num), None)
                if not lote_obj: continue
                
                base_lote = float(lote_obj.monto_base or 0.0)
                ofertas_validas = [(data['monto'], oferente) for oferente, data in ofertas_lote.items() if isinstance(data.get('monto'), (int, float)) and data['monto'] > 0]
                
                if not ofertas_validas: continue

                ofertas_ordenadas = sorted(ofertas_validas)
                top_2 = ofertas_ordenadas[:2]

                analisis_por_lote.append(Paragraph(f"<b><u>Lote {lote_num}: {lote_obj.nombre}</u></b> (Monto Base: RD$ {base_lote:,.2f})", styles['Normal']))
                for i, (monto, oferente) in enumerate(top_2, start=1):
                    dif = monto - base_lote
                    pct = (dif / base_lote * 100) if base_lote > 0 else 0
                    analisis_por_lote.append(Paragraph(f"&nbsp;&nbsp;<b>{i}. {oferente}:</b> RD$ {monto:,.2f} &nbsp;&nbsp;<i>(Diferencia: RD$ {dif:,.2f} / {pct:.2f}%)</i>", styles['Normal']))
                analisis_por_lote.append(Spacer(1, 0.1*inch))
            
            elems.extend(analisis_por_lote)
            elems.append(Spacer(1, 0.3*inch))

            elems.append(Paragraph("Análisis Comparativo (Nuestros Lotes)", styles['h2_left']))
            comparativa_lotes_propios = []
            for lote in sorted([l for l in licitacion.lotes if l.participamos], key=lambda l: l.numero):
                if float(lote.monto_ofertado or 0) <= 0: continue
                
                nuestra_oferta_monto = lote.monto_ofertado
                nuestra_empresa_nombre = f"➡️ {lote.empresa_nuestra or 'Nuestra Oferta'}"
                
                ofertas_competidores = []
                if str(lote.numero) in matriz:
                    ofertas_competidores = [
                        data['monto'] for oferente, data in matriz[str(lote.numero)].items()
                        if oferente != nuestra_empresa_nombre and isinstance(data.get('monto'), (int, float)) and data['monto'] > 0
                    ]

                texto_resultado = f"<b><u>Lote {lote.numero}:</u></b> Nuestra oferta es <b>RD$ {nuestra_oferta_monto:,.2f}</b>. "
                if not ofertas_competidores:
                    texto_resultado += "No hubo otras ofertas válidas de competidores en este lote."
                else:
                    mejor_competidor = min(ofertas_competidores)
                    diferencial = nuestra_oferta_monto - mejor_competidor
                    texto_resultado += f"La mejor oferta competidora fue <b>RD$ {mejor_competidor:,.2f}</b>. Diferencial: <font color='{'red' if diferencial > 0 else 'green'}'><b>RD$ {diferencial:,.2f}</b></font>."
                
                comparativa_lotes_propios.append(Paragraph(texto_resultado, styles['Normal']))
                comparativa_lotes_propios.append(Spacer(1, 0.05*inch))

            if not comparativa_lotes_propios:
                comparativa_lotes_propios.append(Paragraph("No se participó o no se registraron montos de oferta en ningún lote.", styles['Normal']))

            elems.extend(comparativa_lotes_propios)
            
            doc.build(elems)

    def generate_evaluation_report(self, licitacion, resultados_por_lote, file_path):
            """
            Genera un reporte en PDF con los resultados de la evaluación.
            - Horizontal (landscape, carta)
            - Una sección por lote, con el ganador final resaltado.
            """
            doc = SimpleDocTemplate(
                file_path, pagesize=landscape(letter),
                leftMargin=0.5*inch, rightMargin=0.5*inch,
                topMargin=0.9*inch, bottomMargin=0.6*inch
            )

            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(name="LotTitle", parent=styles["Heading3"], spaceAfter=6, textColor=self.GREEN_DARK))
            styles.add(ParagraphStyle(name="Small", parent=styles["BodyText"], fontSize=9, leading=11))
            
            elems = []
            fractions = [0.05, 0.44, 0.08, 0.07, 0.18, 0.08, 0.10]
            col_widths = [doc.width * f for f in fractions]

            for lote_num, resultados_lote in resultados_por_lote.items():
                lote_obj = next((l for l in licitacion.lotes if str(l.numero) == lote_num), None)
                lote_nombre = (lote_obj.nombre if lote_obj else "") or ""
                lot_title = Paragraph(f"Resultados para Lote {lote_num}: {lote_nombre}", styles["LotTitle"])

                header = ["Pos.", "Participante", "Califica", "P. Téc.", "Monto", "P. Eco.", "P. Final"]
                data = [header]

                for i, res in enumerate(resultados_lote, start=1):
                    # --- INICIO CAMBIO: Añadir ícono al ganador ---
                    participante_txt = res['participante']
                    if res.get('es_ganador'):
                        participante_txt = f"🏆 {participante_txt}"
                    # --- FIN CAMBIO ---

                    data.append([
                        i,
                        Paragraph(participante_txt, styles['Small']),
                        "Sí" if res['califica_tecnicamente'] else "NO",
                        f"{res['puntaje_tecnico']:.2f}",
                        f"RD$ {res['monto_ofertado']:,.2f}",
                        f"{res['puntaje_economico']:.2f}",
                        f"{res['puntaje_final']:.2f}",
                    ])

                table = Table(data, hAlign='LEFT', repeatRows=1, colWidths=col_widths)
                
                style = TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), self.GREEN_DARK), ('TEXTCOLOR',(0,0),(-1,0), colors.whitesmoke),
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('BOTTOMPADDING', (0,0), (-1,0), 6),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.black), ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('ALIGN', (1,1), (1,-1), 'LEFT'), ('ALIGN', (4,1), (4,-1), 'RIGHT'),
                    ('FONTNAME', (0,1), (-1,-1), 'Helvetica'), ('FONTSIZE', (0,1), (-1,-1), 9),
                ])

                for i, res in enumerate(resultados_lote, start=1):
                    if i % 2 == 0:
                        style.add('BACKGROUND', (0,i), (-1,i), self.ROW_STRIPE)
                    
                    # --- INICIO CAMBIO: Resaltar la fila del ganador ---
                    if res.get('es_ganador'):
                        style.add('BACKGROUND', (0,i), (-1,i), self.GREEN_LIGHT)
                        style.add('FONTNAME', (0,i), (-1,i), 'Helvetica-Bold')
                    # --- FIN CAMBIO ---

                    if not res['califica_tecnicamente']:
                        style.add('TEXTCOLOR', (0,i), (-1,i), colors.red)

                table.setStyle(style)
                elems.append(KeepTogether([lot_title, Spacer(1, 0.06*inch), table, Spacer(1, 0.25*inch)]))

            doc.build(
                elems,
                onFirstPage=lambda c, d: self._header_footer(c, d, licitacion),
                onLaterPages=lambda c, d: self._header_footer(c, d, licitacion),
            )


    def generate_subsanacion_report(self, licitacion, historial, file_path):
        """Genera un PDF con la tabla del historial de subsanaciones."""
        doc = SimpleDocTemplate(file_path, pagesize=letter,
                                leftMargin=0.7*inch, rightMargin=0.7*inch,
                                topMargin=0.7*inch, bottomMargin=0.7*inch)
        
        styles = getSampleStyleSheet()
        elems = []

        elems.append(Paragraph("Historial de Subsanaciones", styles['h1']))
        elems.append(Paragraph(f"<b>Proceso:</b> {licitacion.numero_proceso} - {licitacion.nombre_proceso}", styles['Normal']))
        elems.append(Paragraph(f"<b>Institución:</b> {licitacion.institucion}", styles['Normal']))
        elems.append(Spacer(1, 0.25*inch))

        header = ["Fecha Solicitud", "Código", "Documento", "Fecha Límite", "Estado"]
        data = [header]

        for row in historial:
            # Extraemos los datos, ignorando el comentario para este reporte simplificado
            fecha_sol, codigo, nombre, fecha_lim, estado, _ = row
            data.append([fecha_sol, codigo, Paragraph(nombre, styles['BodyText']), fecha_lim, estado])

        table = Table(data, hAlign='LEFT', repeatRows=1)
        style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), self.GREEN_DARK),
            ('TEXTCOLOR',(0,0),(-1,0), colors.whitesmoke),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ALIGN', (2,1), (2,-1), 'LEFT'), # Columna del nombre a la izquierda
        ])
        table.setStyle(style)
        
        elems.append(table)
        doc.build(elems)

